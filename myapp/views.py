from django.shortcuts import render, HttpResponse
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from .models import *
from .serializers import *
from rest_framework.response import Response
from rest_framework import status
from datetime import date, datetime
from django.shortcuts import get_object_or_404
from drf_yasg.utils import swagger_auto_schema
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from django.db.models import Max


workbook = openpyxl.Workbook()

sheet1 = workbook.active

data = [
    ['Name', 'product_name', 'Quantity'],
]


file_name = 'data.xlsx'





@swagger_auto_schema(method='GET', responses={200: ProductSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def index(request):
    if request.method == 'GET':
        product = Product.objects.all()
        serialized_products = ProductSerializer(product, many=True)
        return Response(serialized_products.data,    status=status.HTTP_200_OK)


@swagger_auto_schema(method='POST', request_body=ProductSerializer, responses={201: ProductSerializer()})
@api_view(['POST'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def create_product(request):
    if request.method == 'POST':
        create_product = ProductSerializer(data=request.data)
        if create_product.is_valid():
            name = create_product.validated_data.get('name')
            price = create_product.validated_data.get('price')
            expiration_date = create_product.validated_data.get('expiration_date')
            stock = create_product.validated_data.get('stock')
            prod = Product.objects.create(name=name, price=price, expiration_date=expiration_date, stock=stock)
            prod.save()
            return Response(create_product.data, status=status.HTTP_201_CREATED)
        else:
            return Response(create_product.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(method='PUT', request_body=UpdateProductSerializer)
@api_view(['PUT'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def patch_product(request, id):
    if request.method == 'PUT':
        try:
            product = Product.objects.get(id=id)
            update_serializer = UpdateProductSerializer(product, data=request.data, partial=True)
            if update_serializer.is_valid():
                update_serializer.save()
                return Response({"Message": "Product updated"}, status=status.HTTP_200_OK)
            else:
                return Response(update_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": f"Unexpected error {e} occurred."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)



@swagger_auto_schema(method='DELETE', responses={204: ProductSerializer(many=True)})
@api_view(['DELETE'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def product_delete(request, id):
    if request.method == 'DELETE':
        product = get_object_or_404(Product, id=id)
        product.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    else:
        return Response(status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(method='GET', responses={200: ProductSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_product_by_id(request, id):
    if request.method == 'GET':
        product = get_object_or_404(Product, id=id)
        serializer = ProductSerializer(product)
        return Response(serializer.data)

@swagger_auto_schema(method='GET', responses={200: BucketSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_all_busket(request):
    if request.method == 'GET':
        category = Busket.objects.all()
        serialized_products = BucketSerializer(category, many=True)
        return Response(serialized_products.data,    status=status.HTTP_200_OK)

@swagger_auto_schema(method='POST', request_body=BucketSerializer, responses={201: CategorySerializer()})
@api_view(['POST'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def create_busket(request):
    if request.method == 'POST':
        create_category = BucketSerializer(data=request.data)
        if create_category.is_valid():
            owner = create_category.validated_data.get('owner')
            date = create_category.validated_data.get('date')
            prod = Busket.objects.create(owner=owner, date=date)
            prod.save()
            return Response(create_category.data, status=status.HTTP_201_CREATED)
        else:
            return Response(create_category.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(method='PUT', request_body=UpdateBucketSerializer)
@api_view(['PUT'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def patch_busket(request, id):
    if request.method == 'PUT':
        try:
            category = Busket.objects.get(id=id)
            update_serializer = UpdateBucketSerializer(category, data=request.data, partial=True)
            if update_serializer.is_valid():
                update_serializer.save()
                return Response({"Message": "Product updated"}, status=status.HTTP_200_OK)
            else:
                return Response(update_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": f"Unexpected error {e} occurred."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)



@swagger_auto_schema(method='DELETE', responses={204: BucketSerializer(many=True)})
@api_view(['DELETE'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def category_busket(request, id):
    if request.method == 'DELETE':
        category = get_object_or_404(Busket, id=id)
        category.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    else:
        return Response(status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(method='GET', responses={200: BucketSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_busket_by_id(request, id):
    if request.method == 'GET':
        category = get_object_or_404(Busket, id=id)
        serializer = BucketSerializer(category)
        return Response(serializer.data)

@swagger_auto_schema(method='GET', responses={200: CategorySerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_all_category(request):
    if request.method == 'GET':
        category = Category.objects.all()
        serialized_products = CategorySerializer(category, many=True)
        return Response(serialized_products.data,    status=status.HTTP_200_OK)

@swagger_auto_schema(method='POST', request_body=CategorySerializer, responses={201: CategorySerializer()})
@api_view(['POST'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def create_category(request):
    if request.method == 'POST':
        create_category = CategorySerializer(data=request.data)
        if create_category.is_valid():
            name = create_category.validated_data.get('name')
            prod = Product.objects.create(name=name)
            prod.save()
            return Response(create_category.data, status=status.HTTP_201_CREATED)
        else:
            return Response(create_category.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(method='PUT', request_body=CategorySerializer)
@api_view(['PUT'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def patch_category(request, id):
    if request.method == 'PUT':
        try:
            category = Category.objects.get(id=id)
            update_serializer = CategorySerializer(category, data=request.data, partial=True)
            if update_serializer.is_valid():
                update_serializer.save()
                return Response({"Message": "Product updated"}, status=status.HTTP_200_OK)
            else:
                return Response(update_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": f"Unexpected error {e} occurred."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)



@swagger_auto_schema(method='DELETE', responses={204: CategorySerializer(many=True)})
@api_view(['DELETE'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def category_delete(request, id):
    if request.method == 'DELETE':
        category = get_object_or_404(Category, id=id)
        category.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    else:
        return Response(status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(method='GET', responses={200: CategorySerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_category_by_id(request, id):
    if request.method == 'GET':
        category = get_object_or_404(Category, id=id)
        serializer = CategorySerializer(category)
        return Response(serializer.data)


@swagger_auto_schema(method='GET', responses={200: ItemSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_all_items(request):
    if request.method == 'GET':
        category = Item.objects.all()
        serialized_products = ItemSerializer(category, many=True)
        return Response(serialized_products.data,    status=status.HTTP_200_OK)

@swagger_auto_schema(method='POST', request_body=ItemSerializer, responses={201: ItemSerializer()})
@api_view(['POST'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def create_item(request):
    if request.method == 'POST':
        create_category = ItemSerializer(data=request.data)
        if create_category.is_valid():
            busket = create_category.validated_data.get('busket')
            product = create_category.validated_data.get('product')
            quantity = create_category.validated_data.get('quantity')
            prod = Product.objects.create(busket=busket, product=product, quantity=quantity)
            prod.save()
            return Response(create_category.data, status=status.HTTP_201_CREATED)
        else:
            return Response(create_category.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(method='PUT', request_body=ItemSerializer)
@api_view(['PUT'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def patch_item(request, id):
    if request.method == 'PUT':
        try:
            category = Item.objects.get(id=id)
            update_serializer = ItemSerializer(category, data=request.data, partial=True)
            if update_serializer.is_valid():
                update_serializer.save()
                return Response({"Message": "Product updated"}, status=status.HTTP_200_OK)
            else:
                return Response(update_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": f"Unexpected error {e} occurred."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)



@swagger_auto_schema(method='DELETE', responses={204: ItemSerializer(many=True)})
@api_view(['DELETE'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def delete_item(request, id):
    if request.method == 'DELETE':
        category = get_object_or_404(Item, id=id)
        category.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    else:
        return Response(status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(method='GET', responses={200: ItemSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_item_by_id(request, id):
    if request.method == 'GET':
        category = get_object_or_404(Item, id=id)
        serializer = ItemSerializer(category)
        return Response(serializer.data)


@swagger_auto_schema(method='GET', responses={200: CustomerSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_all_customer(request):
    if request.method == 'GET':
        category = Customer.objects.all()
        serialized_products = CustomerSerializer(category, many=True)
        return Response(serialized_products.data,    status=status.HTTP_200_OK)

@swagger_auto_schema(method='POST', request_body=CustomerSerializer, responses={201: CustomerSerializer()})
@api_view(['POST'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def create_customer(request):
    if request.method == 'POST':
        create_category = CustomerSerializer(data=request.data)
        if create_category.is_valid():
            name = create_category.validated_data.get('name')
            email = create_category.validated_data.get('email')
            phone_number = create_category.validated_data.get('phone_number')
            address = create_category.validated_data.get('address')
            prod = Product.objects.create(name=name, email=email, phone_number=phone_number, address=address)
            prod.save()
            return Response(create_category.data, status=status.HTTP_201_CREATED)
        else:
            return Response(create_category.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(method='PUT', request_body=CustomerSerializer)
@api_view(['PUT'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def patch_customer(request, id):
    if request.method == 'PUT':
        try:
            category = Customer.objects.get(id=id)
            update_serializer = CustomerSerializer(category, data=request.data, partial=True)
            if update_serializer.is_valid():
                update_serializer.save()
                return Response({"Message": "Product updated"}, status=status.HTTP_200_OK)
            else:
                return Response(update_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"Error": f"Unexpected error {e} occurred."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"Error": "Invalid request type"}, status=status.HTTP_400_BAD_REQUEST)



@swagger_auto_schema(method='DELETE', responses={204: CustomerSerializer(many=True)})
@api_view(['DELETE'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def delete_customer(request, id):
    if request.method == 'DELETE':
        category = get_object_or_404(Customer, id=id)
        category.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    else:
        return Response(status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(method='GET', responses={200: CustomerSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def get_customer_by_id(request, id):
    if request.method == 'GET':
        category = get_object_or_404(Customer, id=id)
        serializer = CustomerSerializer(category)
        return Response(serializer.data)


@swagger_auto_schema(method='GET', responses={200: ItemSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def user_id_items(request, id):
    customer = Customer.objects.get(id=id)
    print(customer.name)
    bucket = Busket.objects.get(owner=customer)
    print(bucket)
    items = Item.objects.all().filter(busket=bucket)
    print(items)
    res = []
    for item in items:
        res.append([item.busket.owner.name, item.product.name, item.quantity])
        # res += f'{num}. {item.busket}, {item.product}, {item.quantity}(kg/dona/litr).  '
    try:
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        sheet1.append(data[0])
        for row in res:
            sheet1.append(row)
        sheet1['A1'].fill = PatternFill(start_color='000000FF', fill_type='solid')
        sheet1['A1'].fill = PatternFill(fgColor='FFFF00', fill_type='solid')
        workbook.save('user_items.xlsx')

    except Exception as e:
        print(e)
    return HttpResponse(res)




@swagger_auto_schema(method='GET', responses={200: ItemSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def kop_sotilgan(request):
    max = Item.objects.aggregate(Max('quantity'))
    items = Item.objects.get(quantity=max['quantity__max'])
    serializer = ItemSerializer(items)
    res = {"id": serializer.data["id"], "name": items.product.name, "quantity": serializer.data["quantity"], "busket": serializer.data["busket"]}
    return HttpResponse(f'Eng ko\'p sotilgan mahsulot: {res}')

    # return Response(serializer.data)






















@swagger_auto_schema(method='GET', responses={200: ProductSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def summary(request):
    if request.method == 'GET':
        products = Product.objects.all()
        result_all = 0
        bad_products_id = []
        products_id = []
        bad_result = 0
        result = 0
        date = datetime.now()
        for product in products:
            result_all += product.price * product.stock
        for product in products:
            product_id = product.id
            if product.expiration_date < date.date():
                bad_products_id.append(product_id)
            else:
                products_id.append(product_id)

        for id in bad_products_id:
            get_product = Product.objects.get(id=id)
            bad_result += get_product.price * get_product.stock

        for id in products_id:
            get_product = Product.objects.get(id=id)
            result += get_product.price * get_product.stock

        return HttpResponse(f'Barcha mahsulotlar summasi: {result_all}, Muddati o\'tgan mahsulotlar summasi: {bad_result}, Muddati o\'tmagan mahsulotlar summasi: {result}')


@swagger_auto_schema(method='GET', responses={200: ProductSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def all_sell(request, pk):
    if request.method == 'GET':
        res = 0
        data:Busket = Busket.objects.get(id=pk)
        item:Item = Item.objects.all().filter(busket=data)
        for i in item:
            res += i.product.price * i.quantity
        return HttpResponse(f'{data}ning haridlar summasi: {res} so\'m')


@swagger_auto_schema(method='GET', responses={200: ProductSerializer(many=True)})
@api_view(['GET'])
@authentication_classes([BasicAuthentication, SessionAuthentication])
@permission_classes([IsAuthenticated])
def list_of_expired_products(request):
    if request.method == 'GET':
        products = Product.objects.all()
        products_id = []
        date = datetime.now()
        result = ''
        for product in products:
            product_id = product.id
            if product.expiration_date < date.date():
                products_id.append(product_id)
            else:
                pass

        for id in products_id:
            print(id)
            get_product = Product.objects.get(id=id)
            serialized_products = ProductSerializer(get_product)
            print(serialized_products.data)
            result += f'{serialized_products.data}\n'

    return HttpResponse(result)



# Create your views here.
